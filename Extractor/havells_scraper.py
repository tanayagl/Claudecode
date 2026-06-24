#!/usr/bin/env python3
"""
Havells Dealer Locator Scraper
Calls the internal API directly — no browser needed.
Output: havells_dealers.json + havells_dealers.xlsx
"""

import json
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_BASE = "https://havells.com"
HEADERS  = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Referer": "https://havells.com/store-locator",
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest",
}
JSON_FILE  = "havells_dealers.json"
EXCEL_FILE = "havells_dealers.xlsx"

COLS   = ["State", "Name", "Address", "District", "City", "Postal Code", "Mobile", "Alt Mobile", "Email", "Latitude", "Longitude"]
H_FILL = PatternFill("solid", fgColor="1F4E79")
H_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT    = PatternFill("solid", fgColor="D6E4F0")
THIN   = Side(style="thin", color="BFBFBF")
BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WIDTHS = {"State":18,"Name":35,"Address":50,"District":18,"City":20,"Postal Code":12,
          "Mobile":16,"Alt Mobile":16,"Email":35,"Latitude":14,"Longitude":14}


def api_post(path, data):
    r = requests.post(f"{API_BASE}{path}", headers=HEADERS, data=data, timeout=30)
    r.raise_for_status()
    return r.json()


def get_states():
    return api_post("/storelocator/index/getstate", {
        "action": "get_state", "selectWebsite": "Havells", "selectType": "dealer"
    })


def get_dealers(state):
    return api_post("/storelocator/index/getstorelocatorlist", {
        "action": "get_storelocatorlist", "selectWebsite": "Havells",
        "selectType": "dealer", "selectState": state,
        "selectCity": "", "selectCategory": "",
    })


def write_header(ws):
    ws.append(COLS)
    for i in range(1, len(COLS) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = H_FILL; c.font = H_FONT; c.border = BDR
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def write_rows(ws, rows, start=2):
    for i, d in enumerate(rows, start=start):
        fill = ALT if i % 2 == 0 else None
        for j, key in enumerate(COLS, 1):
            c = ws.cell(row=i, column=j, value=d.get(key, ""))
            c.border = BDR
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill


def set_widths(ws):
    for i, h in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(h, 20)


def main():
    print("Fetching states...")
    states = get_states()
    print(f"Found {len(states)} states\n")

    all_dealers = []
    state_data  = {}

    for idx, state in enumerate(states):
        print(f"[{idx+1}/{len(states)}] {state} ... ", end="", flush=True)
        try:
            raw = get_dealers(state)
        except Exception as e:
            print(f"ERROR: {e}")
            raw = []

        dealers = []
        for d in raw:
            addr = " ".join(filter(None, [d.get("address1",""), d.get("address2","")])).strip()
            dealers.append({
                "State":       state,
                "Name":        d.get("firm_name", ""),
                "Address":     addr,
                "District":    d.get("district", ""),
                "City":        d.get("city", ""),
                "Postal Code": d.get("pincode", ""),
                "Mobile":      d.get("mob_number", ""),
                "Alt Mobile":  d.get("mob_number_alt", ""),
                "Email":       d.get("email", ""),
                "Latitude":    d.get("lattitude", ""),
                "Longitude":   d.get("longitude", ""),
            })

        print(f"{len(dealers)} dealers")
        state_data[state] = dealers
        all_dealers.extend(dealers)
        time.sleep(0.3)

    # Save JSON
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(state_data, f, ensure_ascii=False, indent=2)
    print(f"\nJSON → {JSON_FILE}")

    # Save Excel
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Dealers"
    write_header(ws_all)
    write_rows(ws_all, all_dealers)
    set_widths(ws_all)

    for state, dealers in state_data.items():
        if not dealers:
            continue
        ws = wb.create_sheet(title=state[:31])
        write_header(ws)
        write_rows(ws, dealers)
        set_widths(ws)

    wb.save(EXCEL_FILE)
    print(f"Excel → {EXCEL_FILE}")
    print(f"\nDone. Total dealers: {len(all_dealers)}")


if __name__ == "__main__":
    main()
