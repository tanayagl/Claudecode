#!/usr/bin/env python3
"""
Convert havells_dealers.json → havells_dealers.xlsx
Each dealer card: <h5> = Name, <address> lines = Address / District / City / Postal / Tel / Email
"""

import json
import re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "havells_dealers.json"
OUTPUT_FILE = "havells_dealers.xlsx"

HEADERS = ["State", "Name", "Address", "District", "City", "Postal Code", "Phone", "Email"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_dealers(state_name: str, html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    dealers = []

    for card in soup.select("div.bg-light"):
        h5 = card.find("h5")
        name = h5.get_text(strip=True) if h5 else ""
        if not name:
            continue

        addr_el = card.find("address")
        if not addr_el:
            continue

        # Replace <br> tags with newlines
        for br in addr_el.find_all("br"):
            br.replace_with("\n")

        lines = [l.strip() for l in addr_el.get_text().splitlines() if l.strip()]

        address = district = city = postal = phone = email = ""

        for line in lines:
            if line.startswith("District:"):
                district = line.replace("District:", "").strip()
            elif line.startswith("City:"):
                city = line.replace("City:", "").strip()
            elif line.startswith("Postal Code:"):
                postal = line.replace("Postal Code:", "").strip()
            elif line.startswith("Tel:"):
                phone = line.replace("Tel:", "").strip()
            elif line.startswith("Email:"):
                email = line.replace("Email:", "").strip()
            else:
                # First non-labelled line is the street address
                if not address:
                    address = line

        dealers.append({
            "State":       state_name,
            "Name":        name,
            "Address":     address,
            "District":    district,
            "City":        city,
            "Postal Code": postal,
            "Phone":       phone,
            "Email":       email,
        })

    return dealers


def write_header(ws):
    ws.append(HEADERS)
    for col, _ in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def write_rows(ws, dealers, start_row=2):
    for i, d in enumerate(dealers, start=start_row):
        fill = ALT_FILL if i % 2 == 0 else None
        for j, key in enumerate(HEADERS, 1):
            c = ws.cell(row=i, column=j, value=d.get(key, ""))
            c.border    = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill


def auto_width(ws):
    col_widths = {
        "State": 18, "Name": 35, "Address": 50,
        "District": 18, "City": 20, "Postal Code": 12,
        "Phone": 28, "Email": 35,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 20)


def main():
    with open(INPUT_FILE, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Dealers"
    write_header(ws_all)

    total = 0
    for rec in data:
        state   = rec["state_name"]
        dealers = parse_dealers(state, rec["html"])

        if not dealers:
            print(f"  [!] No dealers for {state}")
            continue

        print(f"  {state}: {len(dealers)} dealers")

        # All-dealers sheet
        write_rows(ws_all, dealers, start_row=ws_all.max_row + 1)

        # Per-state sheet
        ws = wb.create_sheet(title=state[:31])
        write_header(ws)
        write_rows(ws, dealers)
        auto_width(ws)

        total += len(dealers)

    auto_width(ws_all)
    wb.save(OUTPUT_FILE)
    print(f"\nTotal: {total} dealers → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
